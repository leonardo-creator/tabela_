from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime
import os
import glob
import pandas as pd
from io import StringIO
from openpyxl import load_workbook
import shutil


class App:

    def _init_(self,login,senha,site, pasta):
        self.login = login
        self.senha = senha
        self.site = site
        self.pasta = pasta
        self.driver = webdriver.Chrome()

    def extrair_dados(self, lista_barragem):

        login  = self.login
        senha = self.senha
        site = self.site

        driver = self.driver
        driver.get(site)

        time.sleep(5)

        elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset/form/ul/li[1]/input') 
        elemento.send_keys(login) 
        time.sleep(1)

        elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset/form/ul/li[2]/input') 
        elemento.send_keys(senha) 
        time.sleep(1)

        elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset/form/ul/li[3]/div') 
        elemento.click()
        time.sleep(1)

        print("conectado com sucesso")

        for barragem in lista_barragem:
            print(f"extraindo dados de: {barragem}")

            elemento = driver.find_element("xpath",f'//option[text()="{barragem}"]')
            elemento.click()
            time.sleep(1)

            elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset[1]/form/input[1]') 
            elemento.send_keys("01/01/2000") 
            time.sleep(1)

            elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset[1]/form/input[2]') 
            elemento.send_keys(datetime.now().strftime("%d/%m/%Y")) 
            time.sleep(1)

            elemento = driver.find_element("xpath",'/html/body/div[2]/fieldset[1]/form/input[3]') 
            elemento.click()
            time.sleep(10)

            elemento = driver.find_element("xpath",'/html/body/div[2]/div[3]/a[1]') 
            elemento.click()
            time.sleep(5)

            elemento = driver.find_element("xpath",'/html/body/div[2]/div[3]/a[2]') 
            elemento.click()
            time.sleep(5)

            download_folder = self.pasta 
            os.chdir(download_folder)

            print("Renomeando arquivos.......")

            for file in glob.glob("*.xls"):
                if file.startswith("Relatorio"):
                    new_file = f"nivel_{barragem}.xls" if "(1)" not in file else f"chuva_{barragem}.xls"
                    os.rename(file, new_file)  # Renomeia o arquivo para o novo nome

                    # Converte o arquivo HTML para xlsx
                    html_path = os.path.join(download_folder, new_file)
                    xlsx_path = os.path.join(download_folder, new_file.replace(".xls", ".xlsx"))
                    self.html_to_xlsx(html_path, xlsx_path)

                    # Remove o arquivo HTML original
                    os.remove(html_path)

        self.renomear_colunas_pasta()
        self.combinar_arquivos_nivel()
        self.combinar_arquivos_chuva()
        self.mesclar_tabelas()
        self.organizar_arquivos()
        self.csvToPowerQuery()

        time.sleep(60)
        driver.quit()

    def html_to_xlsx(self, path, new_path):
        # Ler o arquivo como HTML
        with open(path, "r") as f:
            html_string = f.read()

        # Carrega o arquivo HTML em um dataframe
        data = pd.read_html(StringIO(html_string))[0]

        # Escreve o dataframe em um arquivo xlsx
        data.to_excel(new_path, index=False, engine='openpyxl')

    def renomear_colunas_pasta(self):
        for arquivo in os.listdir(self.pasta):
            if arquivo.endswith('.xlsx') and (arquivo.startswith('chuva_') or arquivo.startswith('nivel_')):
                print(arquivo)
                arquivo_caminho = os.path.join(self.pasta, arquivo)
                wb = load_workbook(arquivo_caminho)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    if arquivo.startswith('nivel_'):
                        sheet['A1'] = "ESTAÇÃO"
                        sheet['B1'] = "DATA/HORA"
                        sheet['C1'] = "NIVEL"
                    elif arquivo.startswith('chuva_'):
                        sheet['A1'] = "ESTAÇÃO"
                        sheet['B1'] = "DATA/HORA"
                        sheet['C1'] = "CHUVA (MM)"
                wb.save(arquivo_caminho)

    def combinar_arquivos_nivel(self):
        all_files = glob.glob(os.path.join(self.pasta, "nivel_*.xlsx"))  
        all_dataframes = []

        for file in all_files:
            print(file)
            df = pd.read_excel(file)
            df['BARRAGEM'] = os.path.basename(file).split('.')[0].replace('nivel_', '')
            print(df)
            all_dataframes.append(df)

        combined_df = pd.concat(all_dataframes, ignore_index=True)
        combined_df = combined_df[['BARRAGEM', 'DATA/HORA', 'NIVEL']]  

        print(combined_df)
        combined_df.to_csv(os.path.join(self.pasta, 'NIVEL_BARRAGEMS.csv'), index=False)
    
    def combinar_arquivos_chuva(self):
        all_files = glob.glob(os.path.join(self.pasta, "chuva_*.xlsx"))  
        all_dataframes = []

        for file in all_files:
            print(file)
            df = pd.read_excel(file)
            df['BARRAGEM'] = os.path.basename(file).split('.')[0].replace('chuva_', '')
            print(df)
            all_dataframes.append(df)

        combined_df = pd.concat(all_dataframes, ignore_index=True)
        combined_df = combined_df[['BARRAGEM', 'DATA/HORA', 'CHUVA (MM)']]  

        print(combined_df)
        combined_df.to_csv(os.path.join(self.pasta, 'CHUVA_BARRAGEMS.csv'), index=False)

    def mesclar_tabelas(self):

        # Carregar os dados
        nivel_barragens = pd.read_csv(os.path.join(self.pasta, 'CHUVA_BARRAGEMS.csv'))
        chuva_barragens = pd.read_csv(os.path.join(self.pasta, 'NIVEL_BARRAGEMS.csv'))

        # Realizar a mescla das tabelas
        tabela_mesclada = pd.merge(nivel_barragens, chuva_barragens, on=['BARRAGEM', 'DATA/HORA'], how='left')

        print(tabela_mesclada)

        # Salvar o resultado em um novo arquivo .xlsx
        tabela_mesclada.to_csv(os.path.join(self.pasta, 'BARRAGENS_CHUVA_NIVEL.csv'), index=False)

    def organizar_arquivos(self):

        self.diretorio_chuvas = os.path.join(self.pasta, "dados chuvas")
        self.diretorio_nivel = os.path.join(self.pasta, "dados nivel")
        
        # Verificar se os diretórios de destino existem, caso contrário, criá-los
        if not os.path.exists(self.diretorio_chuvas):
            os.makedirs(self.diretorio_chuvas)
        if not os.path.exists(self.diretorio_nivel):
            os.makedirs(self.diretorio_nivel)

        # Obter a lista de arquivos na pasta de origem
        arquivos = os.listdir(self.pasta)

        # Mover os arquivos para os diretórios correspondentes
        for arquivo in arquivos:
            if arquivo.lower().startswith("chuva_"):
                shutil.move(os.path.join(self.pasta, arquivo), os.path.join(self.diretorio_chuvas, arquivo))
            elif arquivo.lower().startswith("nivel_"):
                shutil.move(os.path.join(self.pasta, arquivo), os.path.join(self.diretorio_nivel, arquivo))

    def csvToPowerQuery(self):
        df = pd.read_csv(self.pasta + "BARRAGENS_CHUVA_NIVEL.csv")

        # Função para converter para datetime e lidar com exceções
        def try_parsing_date(text):
            try:
                return pd.to_datetime(text, format="%d/%m/%Y %H:%M:%S")
            except ValueError:
                print(f"Não foi possível converter: {text}")
                return None

        # Aplicar a função de conversão em cada valor na coluna
        df['DATA/HORA'] = df['DATA/HORA'].apply(try_parsing_date)

        # Remover linhas que contêm None na coluna 'DATA/HORA'
        df = df.dropna(subset=['DATA/HORA'])

        # Criar coluna ANO
        df['ANO'] = df['DATA/HORA'].dt.year

        # Criar coluna 'MES/DIA'
        df['MES/DIA'] = df['DATA/HORA'].dt.strftime('%m-%d')
        df['MES/DIA'] = pd.to_datetime(df['MES/DIA'] + '-2020', format='%m-%d-%Y')

        # Dicionário para mapear os números dos meses para seus nomes
        mes_map = {
            1: '01 _ Janeiro',
            2: '02 _ Fevereiro',
            3: '03 _ Março',
            4: '04 _ Abril',
            5: '05 _ Maio',
            6: '06 _ Junho',
            7: '07 _ Julho',
            8: '08 _ Agosto',
            9: '09 _ Setembro',
            10: '10 _ Outubro',
            11: '11 _ Novembro',
            12: '12 _ Dezembro',
        }

        # Criar a coluna 'MES' mapeando os números dos meses para seus nomes
        df['MES'] = df['DATA/HORA'].dt.month.map(mes_map)

        # Drop the 'DATA/HORA' column
        df = df.drop(columns=['DATA/HORA'])
        
        # Replace NaNs with 0
        df = df.fillna(0)

        # Agrupando e calculando a média diária
        df = df.groupby(['BARRAGEM', 'ANO', 'MES/DIA', 'MES'], as_index=False).mean()

        # Reformulando o DataFrame para que cada ano tenha suas próprias colunas para CHUVA (MM) e NIVEL
        df_pivot = df.pivot(index=['BARRAGEM', 'MES/DIA', 'MES'], columns='ANO')
        
        # Achatando o MultiIndex das colunas e formatando os nomes das colunas
        df_pivot.columns = ['_'.join(str(i) for i in col) for col in df_pivot.columns]
        
        # Resetando o index para transformar BARRAGEM, MES/DIA e MES em colunas
        df_pivot = df_pivot.reset_index()

        df_pivot.to_csv(self.pasta + "TABELA_POWER_QUERY.csv", decimal=',')
        
        print(df_pivot)



barragens = [
    "Barragem São João",
    "Barragem do Papagaio",
    "Barragem Santo Antônio",
    "Barragem Buritis",
    "Barragem Cocalhinho",
    "Barragem Piaus",
    "Barragem Bananal",
    "Barragem do Coco",
    "Barragem Água Fria",
    "Barragem Campeira",
    "Barragem Horto I",
    "Barragem Carvalhal",
    "Barragem Ribeirão Pinhal",
    "Serra da Natividade",
    "Barragem Urubuzinho",
    "Barragem Fiscal",
    "Palmas ETA 003",
    "Barragem Garrafinha",
    "Barragem Rio Jaguari",
    "Palmas ETA 006",
    "Barragem Pernada",
    "Barragem Zuador",
    "Xinguara",
    "Captação São Borges",
    "Barragem Horto II",
    "Barragem Marcelo",
    "Palmas ETA 007",
    "UTS 02",
    "Operacional 03",
    "Centro de Reservação",
    "ETE Santa Fé",
    "ETE Aureny",
    "Barragem Água Franca"
]

usuario = App("brk","saneatins","", "C:/Users/leonardojuvencio/Downloads","http://hidro.tach.com.br/relatorios.php/")
usuario.extrair_dados(barragens)
